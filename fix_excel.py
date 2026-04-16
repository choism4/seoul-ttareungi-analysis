#!/usr/bin/env python3
"""Excel 수정: VLOOKUP 추가, 콤보차트 보조축 수정"""
import json, os, shutil, zipfile, tempfile
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX = '따릉이_이용패턴_분석.xlsx'
with open('excel_meta.json') as f:
    meta = json.load(f)
with open('seoul_temperature.json') as f:
    temp_data = json.load(f)

DATA_START = meta['DATA_START']
DATA_END = meta['DATA_END']
NAVY = '1B3A5C'
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)

# ── Step 1: Add temperature reference table + VLOOKUP ──
wb = openpyxl.load_workbook(XLSX)

# Add temperature reference table at the bottom of Raw_Data
ws1 = wb['Raw_Data']
TEMP_TABLE_START = DATA_END + 3  # Leave 2 empty rows

ws1.cell(row=TEMP_TABLE_START-1, column=1, value='기온 참조 테이블').font = Font(bold=True, size=11, color=NAVY)
ws1.cell(row=TEMP_TABLE_START, column=9, value='연월')
ws1.cell(row=TEMP_TABLE_START, column=10, value='평균기온(°C)')
ws1.cell(row=TEMP_TABLE_START, column=9).font = HEADER_FONT
ws1.cell(row=TEMP_TABLE_START, column=9).fill = HEADER_FILL
ws1.cell(row=TEMP_TABLE_START, column=10).font = HEADER_FONT
ws1.cell(row=TEMP_TABLE_START, column=10).fill = HEADER_FILL

for i, t in enumerate(temp_data):
    r = TEMP_TABLE_START + 1 + i
    ws1.cell(row=r, column=9, value=t['연월'])
    ws1.cell(row=r, column=10, value=t['평균기온'])
    ws1.cell(row=r, column=9).border = THIN_BORDER
    ws1.cell(row=r, column=10).border = THIN_BORDER
    ws1.cell(row=r, column=10).number_format = '0.0'

TEMP_TABLE_END = TEMP_TABLE_START + len(temp_data)

# Now replace Raw_Data column G (static temperature values) with VLOOKUP formulas
for idx in range(DATA_START, DATA_END + 1):
    ws1.cell(row=idx, column=7, value=
        f'=VLOOKUP(A{idx},$I${TEMP_TABLE_START+1}:$J${TEMP_TABLE_END},2,FALSE)')
    ws1.cell(row=idx, column=7).number_format = '0.0'

print(f"✅ VLOOKUP added: Raw_Data!G{DATA_START}:G{DATA_END} → 기온 참조 테이블 I{TEMP_TABLE_START+1}:J{TEMP_TABLE_END}")

# Also add VLOOKUP in Correlation sheet
ws7 = wb['Correlation']
CORR_START = meta['CORR_START']
CORR_END = meta['CORR_END']
for idx in range(CORR_START, CORR_END + 1):
    ws7.cell(row=idx, column=2, value=
        f'=VLOOKUP(A{idx},Raw_Data!$I${TEMP_TABLE_START+1}:$J${TEMP_TABLE_END},2,FALSE)')
    ws7.cell(row=idx, column=2).number_format = '0.0'

print(f"✅ VLOOKUP added: Correlation!B{CORR_START}:B{CORR_END}")

# Save
wb.save(XLSX)
print("✅ openpyxl save complete")

# ── Step 2: Fix combo chart secondary axis position ──
# Need to modify chart XML directly

tmpdir = tempfile.mkdtemp()
with zipfile.ZipFile(XLSX, 'r') as z:
    z.extractall(tmpdir)

# Find the Dashboard chart (chart5 based on sheet order)
# Charts are in xl/charts/
import glob
chart_files = sorted(glob.glob(f'{tmpdir}/xl/charts/chart*.xml'))
print(f"Chart files: {[os.path.basename(f) for f in chart_files]}")

ns = {'c': 'http://schemas.openxmlformats.org/drawingml/2006/chart',
      'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

for cf in chart_files:
    tree = etree.parse(cf)
    root = tree.getroot()

    # Find combo charts (has both barChart and lineChart)
    bar_charts = root.findall('.//c:barChart', ns)
    line_charts = root.findall('.//c:lineChart', ns)

    if bar_charts and line_charts:
        print(f"  Combo chart found: {os.path.basename(cf)}")
        # Fix secondary value axis position to right
        val_axes = root.findall('.//c:valAx', ns)
        for va in val_axes:
            ax_id = va.find('c:axId', ns)
            ax_pos = va.find('c:axPos', ns)
            crosses = va.find('c:crosses', ns)
            if ax_id is not None and ax_pos is not None:
                aid = ax_id.get('val')
                if crosses is not None and crosses.get('val') == 'max':
                    # This is the secondary axis
                    ax_pos.set('val', 'r')
                    print(f"    Fixed axId={aid} axPos=l→r (secondary axis)")

        tree.write(cf, xml_declaration=True, encoding='UTF-8', standalone=True)

# Re-inject sparklines (they get lost when openpyxl saves)
# Check if sparklines still exist
sheet8_path = f'{tmpdir}/xl/worksheets/sheet8.xml'
with open(sheet8_path, 'r') as f:
    content = f.read()
    has_sparklines = 'sparkline' in content.lower()
    print(f"Sparklines in sheet8 after openpyxl save: {has_sparklines}")

if not has_sparklines:
    print("Re-injecting sparklines...")
    ns_main = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    NS14 = 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main'
    NS_XM = 'http://schemas.microsoft.com/office/excel/2006/main'

    tree_s8 = etree.parse(sheet8_path)
    root_s8 = tree_s8.getroot()

    SPARK_START = meta['SPARK_START']
    years = meta['years']
    age_groups = meta['age_groups']
    data_end_col = chr(64 + len(years) + 1)  # H for 7 years
    spark_col = chr(64 + len(years) + 2)  # I

    sparkline_xml = f'''<ext xmlns:x14="{NS14}" uri="{{05C60535-1F16-4fd2-B633-F4F36F0B64E0}}">
  <x14:sparklineGroups xmlns:xm="{NS_XM}">
    <x14:sparklineGroup type="column" displayEmptyCellsAs="gap">
      <x14:colorSeries rgb="FF4472C4"/>
      <x14:colorNegative rgb="FFFF0000"/>
      <x14:colorAxis rgb="FF000000"/>
      <x14:sparklines>
'''
    for i in range(len(age_groups)):
        data_row = SPARK_START + 1 + i
        sparkline_xml += f'''        <x14:sparkline>
          <xm:f>Dashboard!B{data_row}:{data_end_col}{data_row}</xm:f>
          <xm:sqref>{spark_col}{data_row}</xm:sqref>
        </x14:sparkline>
'''
    sparkline_xml += '''      </x14:sparklines>
    </x14:sparklineGroup>
  </x14:sparklineGroups>
</ext>'''

    extLst = root_s8.find(f'{{{ns_main}}}extLst')
    if extLst is None:
        extLst = etree.SubElement(root_s8, f'{{{ns_main}}}extLst')

    spark_el = etree.fromstring(sparkline_xml)
    extLst.append(spark_el)
    tree_s8.write(sheet8_path, xml_declaration=True, encoding='UTF-8', standalone=True)
    print("✅ Sparklines re-injected")

# Also check/fix pivot table survival
pivot_check = os.path.exists(f'{tmpdir}/xl/pivotTables/pivotTable1.xml')
print(f"Pivot table survived openpyxl save: {pivot_check}")

if not pivot_check:
    print("ERROR: Pivot table XML was removed by openpyxl!")

# Repack
os.remove(XLSX)
with zipfile.ZipFile(XLSX, 'w', zipfile.ZIP_DEFLATED) as zout:
    for root_dir, dirs, files in os.walk(tmpdir):
        for fname in files:
            fpath = os.path.join(root_dir, fname)
            arcname = os.path.relpath(fpath, tmpdir)
            zout.write(fpath, arcname)

shutil.rmtree(tmpdir)
print(f"\n✅ Fixed Excel saved: {XLSX}")

# Final verification
with zipfile.ZipFile(XLSX, 'r') as z:
    names = z.namelist()
    print(f"   Files: {len(names)}")
    print(f"   Pivot: {'xl/pivotTables/pivotTable1.xml' in names}")
    sheet8_data = z.read('xl/worksheets/sheet8.xml').decode('utf-8', errors='ignore')
    print(f"   Sparklines: {'sparkline' in sheet8_data.lower()}")

    # Check VLOOKUP in sheet1
    sheet1_data = z.read('xl/worksheets/sheet1.xml').decode('utf-8', errors='ignore')
    vlookup_count = sheet1_data.lower().count('vlookup')
    print(f"   VLOOKUP formulas in Raw_Data: {vlookup_count}")
