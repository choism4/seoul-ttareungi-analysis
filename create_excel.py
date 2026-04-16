#!/usr/bin/env python3
"""따릉이 분석 Excel 생성 스크립트"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, ScatterChart, PieChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ── Load data ──
with open('monthly_aggregate.json') as f:
    monthly = json.load(f)
with open('seoul_temperature.json') as f:
    temp_data = json.load(f)
with open('gender_yearly.json') as f:
    gender_yearly = json.load(f)
with open('age_yearly.json') as f:
    age_yearly = json.load(f)
with open('age_monthly.json') as f:
    age_monthly = json.load(f)

# Temperature lookup
temp_map = {t['연월']: t['평균기온'] for t in temp_data}

# ── Styles ──
NAVY = '1B3A5C'
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUB_FONT = Font(bold=True, size=11, color='555555')
NUM_FMT = '#,##0'
FLOAT_FMT = '#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)

def style_header(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def style_data_range(ws, r1, r2, c1, c2, fmt=None):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if fmt and c >= 2:
                cell.number_format = fmt

# ══════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1: Raw_Data ──
ws1 = wb.active
ws1.title = 'Raw_Data'
ws1.sheet_properties.tabColor = '4472C4'

ws1['A1'] = '서울시 공공자전거 따릉이 월별 이용 데이터'
ws1['A1'].font = TITLE_FONT
ws1['A2'] = '출처: 서울 열린데이터광장 (data.seoul.go.kr) / 기온: Open-Meteo API'
ws1['A2'].font = SUB_FONT

headers = ['연월', '이용건수', '이동거리(km)', '이용시간(분)', '연도', '월', '평균기온(°C)']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, len(headers))

DATA_START = 4
for idx, m in enumerate(monthly):
    r = DATA_START + idx
    ym = m['연월']
    ws1.cell(row=r, column=1, value=ym)
    ws1.cell(row=r, column=2, value=m['이용건수'])
    ws1.cell(row=r, column=3, value=m['이동거리_km'])
    ws1.cell(row=r, column=4, value=m['이용시간_min'])
    ws1.cell(row=r, column=5, value=m['연도'])
    ws1.cell(row=r, column=6, value=m['월'])
    ws1.cell(row=r, column=7, value=temp_map.get(ym, ''))

DATA_END = DATA_START + len(monthly) - 1  # row 86
style_data_range(ws1, DATA_START, DATA_END, 1, 7, NUM_FMT)
for r in range(DATA_START, DATA_END+1):
    ws1.cell(row=r, column=7).number_format = '0.0'

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 8
ws1.column_dimensions['F'].width = 6
ws1.column_dimensions['G'].width = 14

print(f"Sheet1 Raw_Data: {len(monthly)} rows (row {DATA_START}-{DATA_END})")

# ── Sheet 2: Statistics (기본통계) ──
ws2 = wb.create_sheet('Statistics')
ws2.sheet_properties.tabColor = '70AD47'

ws2['A1'] = '기본 통계 분석'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = '이용건수 / 이동거리 / 이용시간 / 기온 기술통계'
ws2['A2'].font = SUB_FONT

stat_headers = ['통계량', '이용건수', '이동거리(km)', '이용시간(분)', '평균기온(°C)']
for i, h in enumerate(stat_headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(stat_headers))

# B=col2, C=col3, D=col4, G=col7 in Raw_Data
ref_cols = {'B': 'B', 'C': 'C', 'D': 'D', 'G': 'G'}
stat_names = [
    ('평균', 'AVERAGE'), ('최대', 'MAX'), ('최소', 'MIN'),
    ('중앙값', 'MEDIAN'), ('표준편차', 'STDEV'),
    ('최빈값', 'MODE.SNGL'),
    ('상위10%', 'PERCENTILE'), ('하위10%', 'PERCENTILE_LOW'),
    ('상위25%', 'PERCENTILE_Q3'), ('하위25%', 'PERCENTILE_Q1'),
]
col_letters = ['B', 'C', 'D', 'G']  # in Raw_Data
for i, (name, func) in enumerate(stat_names):
    r = 4 + i
    ws2.cell(row=r, column=1, value=name)
    for j, cl in enumerate(col_letters):
        ref = f"Raw_Data!{cl}${DATA_START}:{cl}${DATA_END}"
        if func == 'PERCENTILE':
            formula = f'=PERCENTILE({ref},0.9)'
        elif func == 'PERCENTILE_LOW':
            formula = f'=PERCENTILE({ref},0.1)'
        elif func == 'PERCENTILE_Q3':
            formula = f'=PERCENTILE({ref},0.75)'
        elif func == 'PERCENTILE_Q1':
            formula = f'=PERCENTILE({ref},0.25)'
        else:
            formula = f'={func}({ref})'
        ws2.cell(row=r, column=2+j, value=formula)

style_data_range(ws2, 4, 13, 1, 5, NUM_FMT)
for r in range(4, 14):
    ws2.cell(row=r, column=5).number_format = '0.0'

# Year summary section
ws2['A16'] = '연도별 이용건수 합계'
ws2['A16'].font = Font(bold=True, size=12, color=NAVY)
yr_headers = ['연도', '이용건수', '전년대비증감', '증감률(%)']
for i, h in enumerate(yr_headers, 1):
    ws2.cell(row=17, column=i, value=h)
style_header(ws2, 17, 4)

years = sorted(set(m['연도'] for m in monthly))
for i, yr in enumerate(years):
    r = 18 + i
    ws2.cell(row=r, column=1, value=yr)
    ws2.cell(row=r, column=2, value=f'=SUMPRODUCT((Raw_Data!E${DATA_START}:E${DATA_END}={yr})*Raw_Data!B${DATA_START}:B${DATA_END})')
    if i > 0:
        ws2.cell(row=r, column=3, value=f'=B{r}-B{r-1}')
        ws2.cell(row=r, column=4, value=f'=C{r}/B{r-1}')
    else:
        ws2.cell(row=r, column=3, value='-')
        ws2.cell(row=r, column=4, value='-')

style_data_range(ws2, 18, 18+len(years)-1, 1, 4, NUM_FMT)
for r in range(18, 18+len(years)):
    ws2.cell(row=r, column=4).number_format = PCT_FMT

for c in range(1, 6):
    ws2.column_dimensions[get_column_letter(c)].width = 16

print(f"Sheet2 Statistics: done")

# ── Sheet 3: Pivot_Source ──
ws3 = wb.create_sheet('Pivot_Source')
ws3.sheet_properties.tabColor = 'FFC000'

ws3['A1'] = '피벗테이블 원본 데이터'
ws3['A1'].font = TITLE_FONT
ws3['A2'] = '연령대×연도 이용건수'
ws3['A2'].font = SUB_FONT

# Age groups for pivot
age_groups = ['~10대', '20대', '30대', '40대', '50대', '60대', '70대이상']
pivot_headers = ['연도'] + age_groups + ['합계']
for i, h in enumerate(pivot_headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(pivot_headers))

for i, ay in enumerate(age_yearly):
    r = 4 + i
    ws3.cell(row=r, column=1, value=ay['연도'])
    for j, ag in enumerate(age_groups):
        val = ay.get(ag, 0)
        ws3.cell(row=r, column=2+j, value=val)
    # 합계 formula
    ws3.cell(row=r, column=len(age_groups)+2, value=f'=SUM(B{r}:{get_column_letter(len(age_groups)+1)}{r})')

style_data_range(ws3, 4, 4+len(age_yearly)-1, 1, len(pivot_headers), NUM_FMT)
PIVOT_SRC_END = 4 + len(age_yearly) - 1

# Gender yearly below
GEN_START = PIVOT_SRC_END + 3
ws3.cell(row=GEN_START-1, column=1, value='성별×연도 이용건수').font = Font(bold=True, size=12, color=NAVY)
gen_headers = ['연도', '남성(M)', '여성(F)', '미상', '합계']
for i, h in enumerate(gen_headers, 1):
    ws3.cell(row=GEN_START, column=i, value=h)
style_header(ws3, GEN_START, len(gen_headers))

for i, gy in enumerate(gender_yearly):
    r = GEN_START + 1 + i
    ws3.cell(row=r, column=1, value=gy['연도'])
    ws3.cell(row=r, column=2, value=gy.get('M', 0))
    ws3.cell(row=r, column=3, value=gy.get('F', 0))
    ws3.cell(row=r, column=4, value=gy.get('미상', 0))
    ws3.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})')

style_data_range(ws3, GEN_START+1, GEN_START+len(gender_yearly), 1, 5, NUM_FMT)

for c in range(1, len(pivot_headers)+1):
    ws3.column_dimensions[get_column_letter(c)].width = 14

print(f"Sheet3 Pivot_Source: done")

# ── Sheet 4: CAGR_Ranking ──
ws4 = wb.create_sheet('CAGR_Ranking')
ws4.sheet_properties.tabColor = '7030A0'

ws4['A1'] = '연령대별 CAGR 및 순위'
ws4['A1'].font = TITLE_FONT
ws4['A2'] = 'Compound Annual Growth Rate (2019→2025)'
ws4['A2'].font = SUB_FONT

cagr_headers = ['연령대', '2019 이용건수', '2025 이용건수', '기간(년)', 'CAGR', '순위']
for i, h in enumerate(cagr_headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, len(cagr_headers))

# Data from age_yearly: find 2019 and 2025 values
ay_2019 = next((a for a in age_yearly if a['연도'] == 2019), {})
ay_2025 = next((a for a in age_yearly if a['연도'] == 2025), {})

for i, ag in enumerate(age_groups):
    r = 4 + i
    ws4.cell(row=r, column=1, value=ag)
    v2019 = ay_2019.get(ag, 0)
    v2025 = ay_2025.get(ag, 0)
    ws4.cell(row=r, column=2, value=v2019 if v2019 > 0 else 1)
    ws4.cell(row=r, column=3, value=v2025 if v2025 > 0 else 1)
    ws4.cell(row=r, column=4, value=6)  # 2019→2025 = 6 years
    ws4.cell(row=r, column=5, value=f'=IFERROR(RATE(D{r},0,-B{r},C{r}),0)')
    ws4.cell(row=r, column=6, value=f'=RANK.EQ(E{r},$E$4:$E${4+len(age_groups)-1},0)')

CAGR_END = 4 + len(age_groups) - 1
style_data_range(ws4, 4, CAGR_END, 1, 6, NUM_FMT)
for r in range(4, CAGR_END+1):
    ws4.cell(row=r, column=5).number_format = PCT_FMT

# Chart: CAGR ranking horizontal bar
chart_cagr = BarChart()
chart_cagr.type = 'bar'
chart_cagr.title = '연령대별 CAGR (2019→2025)'
chart_cagr.y_axis.title = '연령대'
chart_cagr.x_axis.title = 'CAGR'
chart_cagr.style = 10
data_ref = Reference(ws4, min_col=5, min_row=3, max_row=CAGR_END)
cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=CAGR_END)
chart_cagr.add_data(data_ref, titles_from_data=True)
chart_cagr.set_categories(cats_ref)
chart_cagr.series[0].graphicalProperties.solidFill = '7030A0'
chart_cagr.width = 20
chart_cagr.height = 12
ws4.add_chart(chart_cagr, 'A12')

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = 16

print(f"Sheet4 CAGR_Ranking: done")

# ── Sheet 5: Moving_Avg_MAE ──
ws5 = wb.create_sheet('Moving_Avg_MAE')
ws5.sheet_properties.tabColor = 'ED7D31'

ws5['A1'] = '이동평균 및 MAE 분석'
ws5['A1'].font = TITLE_FONT
ws5['A2'] = '3개월 / 6개월 이동평균, 평균절대오차'
ws5['A2'].font = SUB_FONT

ma_headers = ['연월', '이용건수', '3개월MA', '6개월MA', '|실제-3MA|', '|실제-6MA|']
for i, h in enumerate(ma_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, len(ma_headers))

MA_START = 4
for idx, m in enumerate(monthly):
    r = MA_START + idx
    ws5.cell(row=r, column=1, value=m['연월'])
    ws5.cell(row=r, column=2, value=f"=Raw_Data!B{DATA_START+idx}")
    # 3-month MA
    if idx >= 2:
        ws5.cell(row=r, column=3, value=f'=AVERAGE(B{r-2}:B{r})')
    # 6-month MA
    if idx >= 5:
        ws5.cell(row=r, column=4, value=f'=AVERAGE(B{r-5}:B{r})')
    # |actual - 3MA|
    if idx >= 3:
        ws5.cell(row=r, column=5, value=f'=IF(C{r}="","",ABS(B{r}-C{r-1}))')
    # |actual - 6MA|
    if idx >= 6:
        ws5.cell(row=r, column=6, value=f'=IF(D{r}="","",ABS(B{r}-D{r-1}))')

MA_END = MA_START + len(monthly) - 1
style_data_range(ws5, MA_START, MA_END, 1, 6, NUM_FMT)

# MAE summary
ws5.cell(row=MA_END+2, column=1, value='MAE (3개월)').font = Font(bold=True)
ws5.cell(row=MA_END+2, column=2, value=f'=AVERAGE(E{MA_START}:E{MA_END})')
ws5.cell(row=MA_END+2, column=2).number_format = NUM_FMT
ws5.cell(row=MA_END+3, column=1, value='MAE (6개월)').font = Font(bold=True)
ws5.cell(row=MA_END+3, column=2, value=f'=AVERAGE(F{MA_START}:F{MA_END})')
ws5.cell(row=MA_END+3, column=2).number_format = NUM_FMT

# Chart: actual vs moving averages
chart_ma = LineChart()
chart_ma.title = '이용건수 실제값 vs 이동평균'
chart_ma.y_axis.title = '이용건수'
chart_ma.x_axis.title = '연월'
chart_ma.style = 10
chart_ma.width = 28
chart_ma.height = 14

for col_idx, label, color in [(2, '실제', '4472C4'), (3, '3개월MA', 'ED7D31'), (4, '6개월MA', '70AD47')]:
    data = Reference(ws5, min_col=col_idx, min_row=3, max_row=MA_END)
    chart_ma.add_data(data, titles_from_data=True)

cats = Reference(ws5, min_col=1, min_row=MA_START, max_row=MA_END)
chart_ma.set_categories(cats)

colors = ['4472C4', 'ED7D31', '70AD47']
for i, c in enumerate(colors):
    chart_ma.series[i].graphicalProperties.line.solidFill = c
    chart_ma.series[i].graphicalProperties.line.width = 20000

ws5.add_chart(chart_ma, 'A' + str(MA_END+6))

for c in range(1, 7):
    ws5.column_dimensions[get_column_letter(c)].width = 14

print(f"Sheet5 Moving_Avg_MAE: done")

# ── Sheet 6: Trend_Seasonal ──
ws6 = wb.create_sheet('Trend_Seasonal')
ws6.sheet_properties.tabColor = 'FF6600'

ws6['A1'] = '추세 분석 및 계절지수'
ws6['A1'].font = TITLE_FONT
ws6['A2'] = 'INTERCEPT/SLOPE 선형추세 + 월별 계절지수 + 2026 예측'
ws6['A2'].font = SUB_FONT

# Time index for SLOPE/INTERCEPT
ws6['A4'] = '선형 추세 파라미터'
ws6['A4'].font = Font(bold=True, size=11)
ws6['A5'] = 'INTERCEPT (절편)'
ws6['B5'] = f'=INTERCEPT(Raw_Data!B${DATA_START}:B${DATA_END},ROW(INDIRECT("1:{len(monthly)}"))-1)'
ws6['B5'].number_format = NUM_FMT
ws6['A6'] = 'SLOPE (기울기)'
ws6['B6'] = f'=SLOPE(Raw_Data!B${DATA_START}:B${DATA_END},ROW(INDIRECT("1:{len(monthly)}"))-1)'
ws6['B6'].number_format = NUM_FMT

# Seasonal index section
ws6['D4'] = '월별 계절지수'
ws6['D4'].font = Font(bold=True, size=11)

si_headers = ['월', '월평균이용건수', '전체평균', '계절지수']
for i, h in enumerate(si_headers):
    ws6.cell(row=5, column=4+i, value=h)
style_header(ws6, 5, 7)

for month in range(1, 13):
    r = 6 + month - 1
    ws6.cell(row=r, column=4, value=month)
    ws6.cell(row=r, column=5, value=f'=AVERAGEIF(Raw_Data!F${DATA_START}:F${DATA_END},{month},Raw_Data!B${DATA_START}:B${DATA_END})')
    ws6.cell(row=r, column=6, value=f'=AVERAGE(Raw_Data!B${DATA_START}:B${DATA_END})')
    ws6.cell(row=r, column=7, value=f'=E{r}/F{r}')
    ws6.cell(row=r, column=5).number_format = NUM_FMT
    ws6.cell(row=r, column=6).number_format = NUM_FMT
    ws6.cell(row=r, column=7).number_format = FLOAT_FMT

style_data_range(ws6, 6, 17, 4, 7)

# Seasonal index bar chart
chart_si = BarChart()
chart_si.title = '월별 계절지수'
chart_si.y_axis.title = '계절지수'
chart_si.x_axis.title = '월'
chart_si.style = 10
data_si = Reference(ws6, min_col=7, min_row=5, max_row=17)
cats_si = Reference(ws6, min_col=4, min_row=6, max_row=17)
chart_si.add_data(data_si, titles_from_data=True)
chart_si.set_categories(cats_si)
chart_si.series[0].graphicalProperties.solidFill = 'FF6600'
chart_si.width = 18
chart_si.height = 10
ws6.add_chart(chart_si, 'D19')

# 2026 forecast section
ws6['A20'] = '2026년 수요 예측 (추세×계절)'
ws6['A20'].font = Font(bold=True, size=12, color=NAVY)

fc_headers = ['월', '시간인덱스', '추세값', '계절지수', '예측이용건수']
for i, h in enumerate(fc_headers):
    ws6.cell(row=21, column=1+i, value=h)
style_header(ws6, 21, 5)

# 2026 = months 85-96 (time index 84-95, since 201901=0)
for month in range(1, 13):
    r = 22 + month - 1
    t_idx = 83 + month  # continuing from 83 months (0-82)
    ws6.cell(row=r, column=1, value=month)
    ws6.cell(row=r, column=2, value=t_idx)
    ws6.cell(row=r, column=3, value=f'=$B$5+$B$6*B{r}')
    ws6.cell(row=r, column=4, value=f'=G{6+month-1}')
    ws6.cell(row=r, column=5, value=f'=C{r}*D{r}')
    ws6.cell(row=r, column=3).number_format = NUM_FMT
    ws6.cell(row=r, column=4).number_format = FLOAT_FMT
    ws6.cell(row=r, column=5).number_format = NUM_FMT

style_data_range(ws6, 22, 33, 1, 5)

for c in range(1, 8):
    ws6.column_dimensions[get_column_letter(c)].width = 16

print(f"Sheet6 Trend_Seasonal: done")

# ── Sheet 7: Correlation ──
ws7 = wb.create_sheet('Correlation')
ws7.sheet_properties.tabColor = 'C00000'

ws7['A1'] = '기온-이용건수 상관분석'
ws7['A1'].font = TITLE_FONT
ws7['A2'] = '산점도 + 추세선 (R² 표시)'
ws7['A2'].font = SUB_FONT

corr_headers = ['연월', '평균기온(°C)', '이용건수']
for i, h in enumerate(corr_headers, 1):
    ws7.cell(row=3, column=i, value=h)
style_header(ws7, 3, 3)

CORR_START = 4
for idx, m in enumerate(monthly):
    r = CORR_START + idx
    ym = m['연월']
    ws7.cell(row=r, column=1, value=ym)
    ws7.cell(row=r, column=2, value=temp_map.get(ym, ''))
    ws7.cell(row=r, column=3, value=m['이용건수'])
    ws7.cell(row=r, column=2).number_format = '0.0'
    ws7.cell(row=r, column=3).number_format = NUM_FMT

CORR_END = CORR_START + len(monthly) - 1

# Correlation coefficient
ws7.cell(row=CORR_END+2, column=1, value='상관계수 (R)').font = Font(bold=True)
ws7.cell(row=CORR_END+2, column=2, value=f'=CORREL(B{CORR_START}:B{CORR_END},C{CORR_START}:C{CORR_END})')
ws7.cell(row=CORR_END+2, column=2).number_format = '0.0000'
ws7.cell(row=CORR_END+3, column=1, value='R² (결정계수)').font = Font(bold=True)
ws7.cell(row=CORR_END+3, column=2, value=f'=RSQ(C{CORR_START}:C{CORR_END},B{CORR_START}:B{CORR_END})')
ws7.cell(row=CORR_END+3, column=2).number_format = '0.0000'
ws7.cell(row=CORR_END+4, column=1, value='SLOPE').font = Font(bold=True)
ws7.cell(row=CORR_END+4, column=2, value=f'=SLOPE(C{CORR_START}:C{CORR_END},B{CORR_START}:B{CORR_END})')
ws7.cell(row=CORR_END+4, column=2).number_format = NUM_FMT
ws7.cell(row=CORR_END+5, column=1, value='INTERCEPT').font = Font(bold=True)
ws7.cell(row=CORR_END+5, column=2, value=f'=INTERCEPT(C{CORR_START}:C{CORR_END},B{CORR_START}:B{CORR_END})')
ws7.cell(row=CORR_END+5, column=2).number_format = NUM_FMT

# Scatter chart
chart_sc = ScatterChart()
chart_sc.title = '기온 vs 이용건수 (산점도 + 추세선)'
chart_sc.x_axis.title = '평균기온(°C)'
chart_sc.y_axis.title = '이용건수'
chart_sc.style = 10
chart_sc.width = 22
chart_sc.height = 14

xvals = Reference(ws7, min_col=2, min_row=CORR_START, max_row=CORR_END)
yvals = Reference(ws7, min_col=3, min_row=CORR_START, max_row=CORR_END)
series = Series(yvals, xvals, title='기온 vs 이용건수')
series.graphicalProperties.line.noFill = True
series.trendline = Trendline(trendlineType='linear', dispRSqr=True, dispEq=True)
chart_sc.series.append(series)

ws7.add_chart(chart_sc, 'E3')

style_data_range(ws7, CORR_START, CORR_END, 1, 3)
for c in range(1, 4):
    ws7.column_dimensions[get_column_letter(c)].width = 14

print(f"Sheet7 Correlation: done")

# ── Sheet 8: Dashboard ──
ws8 = wb.create_sheet('Dashboard')
ws8.sheet_properties.tabColor = '002060'

ws8['A1'] = '따릉이 이용 현황 대시보드'
ws8['A1'].font = Font(bold=True, size=16, color='FFFFFF')
ws8['A1'].fill = PatternFill('solid', fgColor='002060')
ws8.merge_cells('A1:J1')

# Key metrics
ws8['A3'] = '핵심 지표'
ws8['A3'].font = Font(bold=True, size=12, color=NAVY)
metrics = [
    ('총 이용건수 (2019-2025)', f'=SUM(Raw_Data!B{DATA_START}:B{DATA_END})'),
    ('월평균 이용건수', f'=AVERAGE(Raw_Data!B{DATA_START}:B{DATA_END})'),
    ('최대 이용월', f'=INDEX(Raw_Data!A{DATA_START}:A{DATA_END},MATCH(MAX(Raw_Data!B{DATA_START}:B{DATA_END}),Raw_Data!B{DATA_START}:B{DATA_END},0))'),
    ('최대 이용건수', f'=MAX(Raw_Data!B{DATA_START}:B{DATA_END})'),
    ('기온-이용건수 상관계수', f'=Correlation!B{CORR_END+2}'),
]
for i, (label, formula) in enumerate(metrics):
    r = 4 + i
    ws8.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws8.cell(row=r, column=2, value=formula)
    ws8.cell(row=r, column=2).number_format = NUM_FMT

ws8.cell(row=8, column=2).number_format = '0.0000'

# Yearly data for combo chart
ws8['A10'] = '연도별 이용건수 × 평균기온'
ws8['A10'].font = Font(bold=True, size=12, color=NAVY)
combo_headers = ['연도', '이용건수', '이동거리(km)', '평균기온(°C)']
for i, h in enumerate(combo_headers, 1):
    ws8.cell(row=11, column=i, value=h)
style_header(ws8, 11, 4)

for i, yr in enumerate(years):
    r = 12 + i
    ws8.cell(row=r, column=1, value=yr)
    ws8.cell(row=r, column=2, value=f'=SUMPRODUCT((Raw_Data!E${DATA_START}:E${DATA_END}={yr})*Raw_Data!B${DATA_START}:B${DATA_END})')
    ws8.cell(row=r, column=3, value=f'=SUMPRODUCT((Raw_Data!E${DATA_START}:E${DATA_END}={yr})*Raw_Data!C${DATA_START}:C${DATA_END})')
    ws8.cell(row=r, column=4, value=f'=AVERAGEIF(Raw_Data!E${DATA_START}:E${DATA_END},{yr},Raw_Data!G${DATA_START}:G${DATA_END})')
    ws8.cell(row=r, column=2).number_format = NUM_FMT
    ws8.cell(row=r, column=3).number_format = NUM_FMT
    ws8.cell(row=r, column=4).number_format = '0.0'

COMBO_END = 12 + len(years) - 1
style_data_range(ws8, 12, COMBO_END, 1, 4)

# Combo chart: 이용건수 (bar) + 평균기온 (line, secondary axis)
chart_combo = BarChart()
chart_combo.title = '연도별 이용건수 및 평균기온'
chart_combo.y_axis.title = '이용건수'
chart_combo.style = 10
chart_combo.width = 22
chart_combo.height = 14

data_bar = Reference(ws8, min_col=2, min_row=11, max_row=COMBO_END)
chart_combo.add_data(data_bar, titles_from_data=True)
cats_combo = Reference(ws8, min_col=1, min_row=12, max_row=COMBO_END)
chart_combo.set_categories(cats_combo)
chart_combo.series[0].graphicalProperties.solidFill = '4472C4'

# Line overlay for temperature
line_overlay = LineChart()
line_overlay.y_axis.title = '평균기온(°C)'
line_overlay.y_axis.axId = 200
data_line = Reference(ws8, min_col=4, min_row=11, max_row=COMBO_END)
line_overlay.add_data(data_line, titles_from_data=True)
line_overlay.series[0].graphicalProperties.line.solidFill = 'FF0000'
line_overlay.series[0].graphicalProperties.line.width = 25000
line_overlay.y_axis.crosses = 'max'

chart_combo += line_overlay
ws8.add_chart(chart_combo, 'F10')

# Age group pie chart for 2025
ws8['A' + str(COMBO_END+3)] = '2025년 연령대별 이용 비율'
ws8['A' + str(COMBO_END+3)].font = Font(bold=True, size=12, color=NAVY)

pie_start = COMBO_END + 4
pie_headers = ['연령대', '이용건수']
for i, h in enumerate(pie_headers, 1):
    ws8.cell(row=pie_start, column=i, value=h)
style_header(ws8, pie_start, 2)

for i, ag in enumerate(age_groups):
    r = pie_start + 1 + i
    ws8.cell(row=r, column=1, value=ag)
    ws8.cell(row=r, column=2, value=ay_2025.get(ag, 0))
    ws8.cell(row=r, column=2).number_format = NUM_FMT

PIE_END = pie_start + len(age_groups)

chart_pie = PieChart()
chart_pie.title = '2025년 연령대별 이용 비율'
chart_pie.style = 10
chart_pie.width = 16
chart_pie.height = 12
data_pie = Reference(ws8, min_col=2, min_row=pie_start, max_row=PIE_END)
cats_pie = Reference(ws8, min_col=1, min_row=pie_start+1, max_row=PIE_END)
chart_pie.add_data(data_pie, titles_from_data=True)
chart_pie.set_categories(cats_pie)
chart_pie.dataLabels = DataLabelList()
chart_pie.dataLabels.showPercent = True
chart_pie.dataLabels.showCatName = True
ws8.add_chart(chart_pie, 'F' + str(COMBO_END+3))

# Sparkline source data (age groups × years) for OOXML injection later
ws8['A' + str(PIE_END+3)] = '연령대별 연도 추이 (스파크라인용)'
ws8['A' + str(PIE_END+3)].font = Font(bold=True, size=12, color=NAVY)
spark_start = PIE_END + 4
spark_headers = ['연령대'] + [str(y) for y in years]
for i, h in enumerate(spark_headers, 1):
    ws8.cell(row=spark_start, column=i, value=h)
style_header(ws8, spark_start, len(spark_headers))

for i, ag in enumerate(age_groups):
    r = spark_start + 1 + i
    ws8.cell(row=r, column=1, value=ag)
    for j, yr_data in enumerate(age_yearly):
        ws8.cell(row=r, column=2+j, value=yr_data.get(ag, 0))
        ws8.cell(row=r, column=2+j).number_format = NUM_FMT

SPARK_END = spark_start + len(age_groups)
# Add sparkline column header
ws8.cell(row=spark_start, column=len(spark_headers)+1, value='추이')
ws8.cell(row=spark_start, column=len(spark_headers)+1).font = HEADER_FONT
ws8.cell(row=spark_start, column=len(spark_headers)+1).fill = HEADER_FILL

for c in range(1, 12):
    ws8.column_dimensions[get_column_letter(c)].width = 16

print(f"Sheet8 Dashboard: done")

# ── Save ──
OUTPUT = '따릉이_이용패턴_분석.xlsx'
wb.save(OUTPUT)
print(f"\n✅ Excel saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")

# Store metadata for OOXML injection
meta = {
    'DATA_START': DATA_START, 'DATA_END': DATA_END,
    'CORR_START': CORR_START, 'CORR_END': CORR_END,
    'MA_START': MA_START, 'MA_END': MA_END,
    'CAGR_END': CAGR_END, 'COMBO_END': COMBO_END,
    'PIE_END': PIE_END, 'SPARK_START': spark_start, 'SPARK_END': SPARK_END,
    'GEN_START': GEN_START,
    'PIVOT_SRC_END': PIVOT_SRC_END,
    'years': years,
    'age_groups': age_groups,
    'sheet_names': wb.sheetnames,
}
with open('excel_meta.json', 'w') as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)
print(f"   Meta saved: excel_meta.json")
