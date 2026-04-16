#!/usr/bin/env python3
"""V2 신규 시트 추가 — 4인 전문가 비평 반영.

실행 순서: create_excel.py → add_v2_sheets.py (이 파일) → fix_excel.py → inject_ooxml.py

추가 시트:
- Residual_Analysis  (기온 효과 제거 후 구조적 감소 — 킬러)
- LINEST_Multi       (다중회귀로 기온 순효과 분리)
- Forecast_Benchmark (Seasonal Naive 포함 4모델 + 2025 hold-out + 신뢰구간)
- District_Analysis  (25자치구 × 지니계수)
- Unit_Economics     (대당 회전율)
- Misu_Adjustment    (미상 안분 보정 CAGR)

기존 시트 보강 없음 (기존 시트는 그대로 유지).
"""
import json
import os
from collections import defaultdict

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX = '따릉이_이용패턴_분석.xlsx'
wb = openpyxl.load_workbook(XLSX)

with open('excel_meta.json') as f:
    meta = json.load(f)
with open('monthly_aggregate.json') as f:
    monthly = json.load(f)
with open('age_yearly.json') as f:
    age_yearly = json.load(f)
with open('gender_yearly.json') as f:
    gender_yearly = json.load(f)
with open('district_monthly.json') as f:
    district_monthly = json.load(f)
with open('external_context.json') as f:
    ctx = json.load(f)

DATA_START = meta['DATA_START']
DATA_END = meta['DATA_END']
N = DATA_END - DATA_START + 1

NAVY = '1B3A5C'
ORANGE = 'ED7D31'
RED = 'C00000'
GREEN = '70AD47'
PURPLE = '7030A0'
BLUE = '4472C4'

HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUB_FONT = Font(bold=True, size=11, color='555555')
BOLD_FONT = Font(bold=True)
NUM_FMT = '#,##0'
FLOAT_FMT = '#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def style_data_range(ws, r1, r2, c1, c2, fmt=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if fmt and c >= 2:
                cell.number_format = fmt


# ═══════════════════════════════════════════════════════════════
# Sheet 9: Residual_Analysis (킬러 인사이트)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Residual_Analysis')
ws.sheet_properties.tabColor = RED

ws['A1'] = '잔차 분석 — 기온 효과 제거 후 구조적 감소'
ws['A1'].font = TITLE_FONT
ws['A2'] = '단순회귀(이용~기온) 잔차의 시계열 및 Durbin-Watson 통계량'
ws['A2'].font = SUB_FONT

headers = ['연월', '이용건수', '평균기온(°C)', '예측값(단순회귀)', '잔차', '잔차^2']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

RES_START = 4
RES_END = RES_START + N - 1

# SLOPE/INTERCEPT from Correlation sheet (already computed there)
# Re-derive here for clarity
for idx in range(N):
    r = RES_START + idx
    raw_r = DATA_START + idx
    ws.cell(row=r, column=1, value=f'=Raw_Data!A{raw_r}')
    ws.cell(row=r, column=2, value=f'=Raw_Data!B{raw_r}')
    ws.cell(row=r, column=3, value=f'=Raw_Data!G{raw_r}')
    # 예측값 = INTERCEPT + SLOPE * 기온
    ws.cell(row=r, column=4, value=(
        f'=INTERCEPT(Raw_Data!B${DATA_START}:B${DATA_END},Raw_Data!G${DATA_START}:G${DATA_END})'
        f'+SLOPE(Raw_Data!B${DATA_START}:B${DATA_END},Raw_Data!G${DATA_START}:G${DATA_END})*C{r}'
    ))
    ws.cell(row=r, column=5, value=f'=B{r}-D{r}')  # 잔차
    ws.cell(row=r, column=6, value=f'=E{r}^2')
    ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.cell(row=r, column=3).number_format = '0.0'
    ws.cell(row=r, column=4).number_format = NUM_FMT
    ws.cell(row=r, column=5).number_format = NUM_FMT
    ws.cell(row=r, column=6).number_format = NUM_FMT

style_data_range(ws, RES_START, RES_END, 1, 6)

# 통계 요약
stat_row = RES_END + 2
ws.cell(row=stat_row, column=1, value='Durbin-Watson').font = BOLD_FONT
ws.cell(row=stat_row, column=2, value=(
    f'=SUMXMY2(E{RES_START + 1}:E{RES_END},E{RES_START}:E{RES_END - 1})'
    f'/SUMSQ(E{RES_START}:E{RES_END})'
))
ws.cell(row=stat_row, column=2).number_format = FLOAT_FMT
ws.cell(row=stat_row, column=3,
        value='(2 근사 = 자기상관 없음 / <1.5 양의 자기상관 / >2.5 음의 자기상관)').font = SUB_FONT

ws.cell(row=stat_row + 1, column=1, value='잔차 RMSE').font = BOLD_FONT
ws.cell(row=stat_row + 1, column=2,
        value=f'=SQRT(AVERAGE(F{RES_START}:F{RES_END}))')
ws.cell(row=stat_row + 1, column=2).number_format = NUM_FMT

ws.cell(row=stat_row + 2, column=1, value='잔차 vs 시간인덱스 상관(r)').font = BOLD_FONT
# 시간 인덱스: ROW() - RES_START
ws.cell(row=stat_row + 2, column=2, value=(
    f'=CORREL(E{RES_START}:E{RES_END},ROW(E{RES_START}:E{RES_END}))'
))
ws.cell(row=stat_row + 2, column=2).number_format = FLOAT_FMT
ws.cell(row=stat_row + 2, column=3,
        value='(음의 상관 = 시간이 갈수록 잔차 감소 = 구조적 감소 징후)').font = SUB_FONT

ws.cell(row=stat_row + 3, column=1, value='2023년 이후 평균 잔차').font = BOLD_FONT
# 2023년 = monthly index 49 (idx 48, r = RES_START + 48)
r_2023 = RES_START + 48
ws.cell(row=stat_row + 3, column=2, value=f'=AVERAGE(E{r_2023}:E{RES_END})')
ws.cell(row=stat_row + 3, column=2).number_format = NUM_FMT
ws.cell(row=stat_row + 3, column=3,
        value='(음수 = 기온 대비 이용량이 기대치를 밑돎 = 구조적 감소 증거)').font = SUB_FONT

# 잔차 시계열 차트
chart_res = LineChart()
chart_res.title = '잔차 시계열 — 2023년 이후 구조적 감소'
chart_res.y_axis.title = '잔차 (실제 - 기온 기반 예측)'
chart_res.x_axis.title = '연월'
chart_res.style = 10
chart_res.width = 28
chart_res.height = 12
data_res = Reference(ws, min_col=5, min_row=3, max_row=RES_END)
cats_res = Reference(ws, min_col=1, min_row=RES_START, max_row=RES_END)
chart_res.add_data(data_res, titles_from_data=True)
chart_res.set_categories(cats_res)
chart_res.series[0].graphicalProperties.line.solidFill = RED
chart_res.series[0].graphicalProperties.line.width = 24000
ws.add_chart(chart_res, f'A{stat_row + 6}')

for c in range(1, 7):
    ws.column_dimensions[get_column_letter(c)].width = 16
ws.column_dimensions['C'].width = 60  # 설명 칼럼

print('✓ Residual_Analysis 추가')

# ═══════════════════════════════════════════════════════════════
# Sheet 10: LINEST_Multi (다중회귀)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('LINEST_Multi')
ws.sheet_properties.tabColor = PURPLE

ws['A1'] = '다중회귀 분석 — 기온 순효과 분리'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'LINEST 배열수식: 이용건수 ~ 기온 + 시간인덱스 + COVID더미 + 월더미'
ws['A2'].font = SUB_FONT

# X 행렬 (설계 행렬) 을 Raw_Data에서 참조해 이 시트에 구성
# Columns: B=y, C=Temp, D=t, E=D_covid, F~P=M2..M12 (기저=1월)
x_headers = ['연월', 'y(이용건수)', 'Temp', 't', 'D_covid']
month_cols = [f'M{m}' for m in range(2, 13)]
x_headers += month_cols
for i, h in enumerate(x_headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(x_headers))

X_START = 4
for idx in range(N):
    r = X_START + idx
    raw_r = DATA_START + idx
    ws.cell(row=r, column=1, value=f'=Raw_Data!A{raw_r}')
    ws.cell(row=r, column=2, value=f'=Raw_Data!B{raw_r}')
    ws.cell(row=r, column=3, value=f'=Raw_Data!G{raw_r}')  # Temp
    ws.cell(row=r, column=4, value=idx)  # t = 0..N-1
    # COVID dummy: 2020.03~2022.04 (index 14~40 for 2019.01 start)
    ym = monthly[idx]['연월']
    covid = 1 if ('202003' <= str(ym) <= '202204') else 0
    ws.cell(row=r, column=5, value=covid)
    # month dummies (M2..M12)
    m = monthly[idx]['월']
    for j, mm in enumerate(range(2, 13)):
        ws.cell(row=r, column=6 + j, value=1 if m == mm else 0)

X_END = X_START + N - 1
style_data_range(ws, X_START, X_END, 1, len(x_headers))

# LINEST 배열수식
# X 범위: C~P = 14개 독립변수
X_FIRST_COL = 'C'
X_LAST_COL = get_column_letter(5 + len(month_cols))  # E(covid) + 11 months = col 16 = P
NUM_VARS = 2 + 1 + 11  # Temp + t + covid + 11 month dummies = 14

linest_start_row = X_END + 3
ws.cell(row=linest_start_row, column=1, value='LINEST 다중회귀 결과').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=linest_start_row + 1, column=1, value='(배열수식 — 상위 5행 × 15열 스필)').font = SUB_FONT

# LINEST가 역순으로 coefficient를 뱉음: [β_last, ..., β_1, β_0]
# 스필 범위: col A(label) + coefs in order
ws.cell(row=linest_start_row + 3, column=1, value='LINEST 블록 (5행×15열 — β, SE, R²/SE_est, F/df, SS_reg/SS_res)').font = BOLD_FONT

linest_block_row = linest_start_row + 4
formula = (
    f'=LINEST(B{X_START}:B{X_END},'
    f'{X_FIRST_COL}{X_START}:{X_LAST_COL}{X_END},'
    f'TRUE,TRUE)'
)
ws.cell(row=linest_block_row, column=1, value=formula)
# openpyxl array formula support: use ws.formula_attributes
from openpyxl.worksheet.formula import ArrayFormula
array_range = f'A{linest_block_row}:{get_column_letter(NUM_VARS + 1)}{linest_block_row + 4}'
ws[f'A{linest_block_row}'] = ArrayFormula(array_range, formula)

# 해석 가이드
guide_row = linest_block_row + 7
ws.cell(row=guide_row, column=1, value='해석 가이드 (LINEST 역순 규칙)').font = Font(bold=True, size=11, color=NAVY)
guide = [
    '• 1행: 계수 β — 우측부터 역순 → [β_M12, β_M11, ..., β_M2, β_covid, β_t, β_temp, β_intercept]',
    '• 2행: 표준오차 SE (t = β/SE, |t|>2 → 유의)',
    '• 3행: [0,0]=R², [0,1]=Y 추정 표준오차 (그 외 셀=#N/A 정상)',
    '• 4행: [0,0]=F 통계량, [0,1]=잔차 자유도',
    '• 5행: [0,0]=회귀 SS, [0,1]=잔차 SS',
    '',
    '※ 단순회귀의 기온 계수와 다중회귀의 기온 계수를 비교해',
    '  "기온 외 변수(시간·COVID·계절) 통제 후 기온 순효과"를 해석할 것.',
]
for i, g in enumerate(guide):
    ws.cell(row=guide_row + 1 + i, column=1, value=g).font = SUB_FONT
    ws.merge_cells(start_row=guide_row + 1 + i, start_column=1, end_row=guide_row + 1 + i, end_column=10)

# 단순회귀(기온만) 비교
cmp_row = guide_row + 12
ws.cell(row=cmp_row, column=1, value='비교: 단순회귀 vs 다중회귀 기온 계수').font = Font(bold=True, size=11, color=NAVY)
ws.cell(row=cmp_row + 1, column=1, value='단순회귀 SLOPE (기온만)').font = BOLD_FONT
ws.cell(row=cmp_row + 1, column=2,
        value=f'=SLOPE(Raw_Data!B${DATA_START}:B${DATA_END},Raw_Data!G${DATA_START}:G${DATA_END})')
ws.cell(row=cmp_row + 1, column=2).number_format = NUM_FMT
ws.cell(row=cmp_row + 2, column=1, value='다중회귀 β_temp (LINEST 블록 참조)').font = BOLD_FONT
# β_temp은 역순에서 2번째 뒤 (intercept 바로 앞): 15열 중 col N-1 = col index 14 (0-indexed 13)
# 사실 LINEST 스필 블록에서 β_temp는 B{linest_block_row} (intercept 다음 역순 첫 계수)
ws.cell(row=cmp_row + 2, column=2,
        value=f'=INDEX(A{linest_block_row}:{get_column_letter(NUM_VARS + 1)}{linest_block_row},1,{NUM_VARS})')
ws.cell(row=cmp_row + 2, column=2).number_format = NUM_FMT
ws.cell(row=cmp_row + 3, column=1, value='R² (LINEST)').font = BOLD_FONT
ws.cell(row=cmp_row + 3, column=2,
        value=f'=INDEX(A{linest_block_row + 2}:B{linest_block_row + 2},1,1)')
ws.cell(row=cmp_row + 3, column=2).number_format = FLOAT_FMT

for c in range(1, 18):
    ws.column_dimensions[get_column_letter(c)].width = 12
ws.column_dimensions['A'].width = 36

print('✓ LINEST_Multi 추가')

# ═══════════════════════════════════════════════════════════════
# Sheet 11: Forecast_Benchmark
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Forecast_Benchmark')
ws.sheet_properties.tabColor = ORANGE

ws['A1'] = '예측 모델 벤치마크 — 4모델 비교 (2025 hold-out)'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Naive / Seasonal Naive / 3개월MA / 추세×계절지수 — MAE · MAPE · RMSE'
ws['A2'].font = SUB_FONT

fb_headers = ['연월', '이용건수', 'Naive(t-1)', 'SeasonalNaive(t-12)', '3MA',
              '추세×계절', '|Naive|오차', '|SN|오차', '|3MA|오차', '|T×S|오차']
for i, h in enumerate(fb_headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(fb_headers))

FB_START = 4
FB_END = FB_START + N - 1

# 추세×계절: Trend_Seasonal 시트의 INTERCEPT/SLOPE + 계절지수 재활용
# 계절지수는 ws['Trend_Seasonal']의 G6:G17 (month 1~12의 계절지수)
for idx in range(N):
    r = FB_START + idx
    raw_r = DATA_START + idx
    m = monthly[idx]['월']
    ws.cell(row=r, column=1, value=f'=Raw_Data!A{raw_r}')
    ws.cell(row=r, column=2, value=f'=Raw_Data!B{raw_r}')
    # Naive: yₜ₋₁
    if idx >= 1:
        ws.cell(row=r, column=3, value=f'=B{r - 1}')
    # Seasonal Naive: yₜ₋₁₂
    if idx >= 12:
        ws.cell(row=r, column=4, value=f'=B{r - 12}')
    # 3MA (trailing)
    if idx >= 3:
        ws.cell(row=r, column=5, value=f'=AVERAGE(B{r - 3}:B{r - 1})')
    # 추세×계절: Trend_Seasonal!$B$5 + Trend_Seasonal!$B$6 * idx * 계절지수
    ws.cell(row=r, column=6, value=(
        f'=(Trend_Seasonal!$B$5+Trend_Seasonal!$B$6*{idx})'
        f'*INDEX(Trend_Seasonal!$G$6:$G$17,{m})'
    ))
    # 절대오차
    for col_pred, col_abs in [(3, 7), (4, 8), (5, 9), (6, 10)]:
        ws.cell(row=r, column=col_abs, value=(
            f'=IF({get_column_letter(col_pred)}{r}="","",ABS(B{r}-{get_column_letter(col_pred)}{r}))'
        ))

style_data_range(ws, FB_START, FB_END, 1, len(fb_headers), NUM_FMT)

# Hold-out: 2025년만 (마지막 12개월) → FB_END-11 ~ FB_END
HO_START = FB_END - 11
HO_END = FB_END

summary_row = FB_END + 2
ws.cell(row=summary_row, column=1, value='Hold-out (2025): 모델별 오차 지표').font = Font(bold=True, size=12, color=NAVY)

sum_headers = ['모델', 'MAE', 'MAPE', 'RMSE']
for i, h in enumerate(sum_headers, 1):
    ws.cell(row=summary_row + 1, column=i, value=h)
style_header(ws, summary_row + 1, 4)

models = [
    ('Naive(t-1)', 'G', 'C'),
    ('Seasonal Naive(t-12)', 'H', 'D'),
    ('3개월 MA', 'I', 'E'),
    ('추세×계절', 'J', 'F'),
]
for i, (name, abs_col, pred_col) in enumerate(models):
    r = summary_row + 2 + i
    ws.cell(row=r, column=1, value=name).font = BOLD_FONT
    ws.cell(row=r, column=2, value=f'=AVERAGE({abs_col}{HO_START}:{abs_col}{HO_END})')
    ws.cell(row=r, column=2).number_format = NUM_FMT
    # MAPE: |err|/y 평균
    ws.cell(row=r, column=3, value=(
        f'=AVERAGE(IF(B{HO_START}:B{HO_END}>0,{abs_col}{HO_START}:{abs_col}{HO_END}/B{HO_START}:B{HO_END},NA()))'
    ))
    ws[f'C{r}'] = openpyxl.worksheet.formula.ArrayFormula(
        f'C{r}',
        f'=AVERAGE(IF(B{HO_START}:B{HO_END}>0,{abs_col}{HO_START}:{abs_col}{HO_END}/B{HO_START}:B{HO_END},NA()))'
    )
    ws.cell(row=r, column=3).number_format = PCT_FMT
    # RMSE: sqrt(mean(err^2))
    ws.cell(row=r, column=4, value=(
        f'=SQRT(SUMSQ({abs_col}{HO_START}:{abs_col}{HO_END})/COUNT({abs_col}{HO_START}:{abs_col}{HO_END}))'
    ))
    ws.cell(row=r, column=4).number_format = NUM_FMT

style_data_range(ws, summary_row + 2, summary_row + 5, 1, 4)

# 2026 예측 + 신뢰구간 (추세×계절 기준)
fc_row = summary_row + 8
ws.cell(row=fc_row, column=1, value='2026년 예측 — 추세×계절 모델 (95% 신뢰구간 ±1.96·RMSE)').font = Font(bold=True, size=12, color=NAVY)

fc_headers = ['월', '시간인덱스', '예측', '하한(-1.96RMSE)', '상한(+1.96RMSE)']
for i, h in enumerate(fc_headers, 1):
    ws.cell(row=fc_row + 1, column=i, value=h)
style_header(ws, fc_row + 1, len(fc_headers))

for m in range(1, 13):
    r = fc_row + 1 + m
    t_idx = N + m - 1  # 계속: 2026.1 = 83, 2026.12 = 94
    ws.cell(row=r, column=1, value=m)
    ws.cell(row=r, column=2, value=t_idx)
    ws.cell(row=r, column=3, value=(
        f'=(Trend_Seasonal!$B$5+Trend_Seasonal!$B$6*B{r})'
        f'*INDEX(Trend_Seasonal!$G$6:$G$17,A{r})'
    ))
    # RMSE of 추세×계절 on hold-out = D{summary_row+5}
    rmse_cell = f'D{summary_row + 5}'
    ws.cell(row=r, column=4, value=f'=C{r}-1.96*{rmse_cell}')
    ws.cell(row=r, column=5, value=f'=C{r}+1.96*{rmse_cell}')
    for c in range(3, 6):
        ws.cell(row=r, column=c).number_format = NUM_FMT

style_data_range(ws, fc_row + 2, fc_row + 13, 1, 5)

# 차트: 실제값 + 2026 예측 + 신뢰구간
chart_fb = LineChart()
chart_fb.title = '2026 예측 + 95% 신뢰구간'
chart_fb.y_axis.title = '이용건수'
chart_fb.x_axis.title = '월'
chart_fb.style = 10
chart_fb.width = 24
chart_fb.height = 12
data_fc = Reference(ws, min_col=3, min_row=fc_row + 1, max_row=fc_row + 13)
chart_fb.add_data(data_fc, titles_from_data=True)
data_lo = Reference(ws, min_col=4, min_row=fc_row + 1, max_row=fc_row + 13)
chart_fb.add_data(data_lo, titles_from_data=True)
data_hi = Reference(ws, min_col=5, min_row=fc_row + 1, max_row=fc_row + 13)
chart_fb.add_data(data_hi, titles_from_data=True)
cats_fc = Reference(ws, min_col=1, min_row=fc_row + 2, max_row=fc_row + 13)
chart_fb.set_categories(cats_fc)
chart_fb.series[0].graphicalProperties.line.solidFill = BLUE
chart_fb.series[0].graphicalProperties.line.width = 28000
chart_fb.series[1].graphicalProperties.line.solidFill = 'BBBBBB'
chart_fb.series[1].graphicalProperties.line.dashStyle = 'dash'
chart_fb.series[2].graphicalProperties.line.solidFill = 'BBBBBB'
chart_fb.series[2].graphicalProperties.line.dashStyle = 'dash'
ws.add_chart(chart_fb, f'G{fc_row + 1}')

for c in range(1, 11):
    ws.column_dimensions[get_column_letter(c)].width = 14

print('✓ Forecast_Benchmark 추가')

# ═══════════════════════════════════════════════════════════════
# Sheet 12: District_Analysis
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('District_Analysis')
ws.sheet_properties.tabColor = GREEN

ws['A1'] = '자치구별 이용 분포 — 공간 형평성'
ws['A1'].font = TITLE_FONT
ws['A2'] = '25자치구 이용건수 · 인구 · 1인당 이용 · 지니계수'
ws['A2'].font = SUB_FONT

# 자치구별 2025 이용건수 계산
gu_2025 = defaultdict(int)
for row in district_monthly:
    if str(row['ym']).startswith('2025') and row['자치구'] != '미상':
        gu_2025[row['자치구']] += row['이용건수']

# 서울 자치구 인구 (2025-01 기준, 서울통계)
# 출처: 서울시 주민등록인구통계 (단위: 명, 반올림)
population = {
    '강남구': 558_000, '강동구': 460_000, '강북구': 284_000, '강서구': 553_000,
    '관악구': 487_000, '광진구': 340_000, '구로구': 393_000, '금천구': 236_000,
    '노원구': 495_000, '도봉구': 306_000, '동대문구': 342_000, '동작구': 382_000,
    '마포구': 369_000, '서대문구': 303_000, '서초구': 413_000, '성동구': 282_000,
    '성북구': 428_000, '송파구': 651_000, '양천구': 440_000, '영등포구': 375_000,
    '용산구': 213_000, '은평구': 469_000, '종로구': 141_000, '중구': 121_000,
    '중랑구': 386_000,
}

# 자치구별 대여소 수
master = __import__('pandas').read_csv('station_master.csv')
station_count = master['자치구'].value_counts().to_dict()

headers = ['자치구', '2025 이용건수', '인구(명)', '1인당 이용(회/년)', '대여소 수', '대여소당 이용']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

gus = sorted(gu_2025.keys())
D_START = 4
for i, gu in enumerate(gus):
    r = D_START + i
    pop = population.get(gu, 1)
    scount = station_count.get(gu, 1)
    usage = gu_2025[gu]
    ws.cell(row=r, column=1, value=gu)
    ws.cell(row=r, column=2, value=usage)
    ws.cell(row=r, column=3, value=pop)
    ws.cell(row=r, column=4, value=f'=B{r}/C{r}')
    ws.cell(row=r, column=5, value=scount)
    ws.cell(row=r, column=6, value=f'=B{r}/E{r}')
    ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.cell(row=r, column=3).number_format = NUM_FMT
    ws.cell(row=r, column=4).number_format = FLOAT_FMT
    ws.cell(row=r, column=5).number_format = NUM_FMT
    ws.cell(row=r, column=6).number_format = NUM_FMT

D_END = D_START + len(gus) - 1
style_data_range(ws, D_START, D_END, 1, 6)

# 지니계수 계산 (1인당 이용 기준)
gini_row = D_END + 2
ws.cell(row=gini_row, column=1, value='자치구 1인당 이용의 지니계수').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=gini_row + 1, column=1, value='n').font = BOLD_FONT
ws.cell(row=gini_row + 1, column=2, value=len(gus))

ws.cell(row=gini_row + 2, column=1, value='평균 1인당 이용').font = BOLD_FONT
ws.cell(row=gini_row + 2, column=2, value=f'=AVERAGE(D{D_START}:D{D_END})')
ws.cell(row=gini_row + 2, column=2).number_format = FLOAT_FMT

# Gini = (1 / (2 * n² * μ)) * ΣΣ |x_i - x_j|
# Excel에서는 SUMPRODUCT로 구현 — 보조 시트 없이 한 셀로 계산하려면
# 모든 페어의 절대차 합이 필요. 가장 간단한 공식: 정렬 후 2i-n-1 가중합.
# 정렬된 값이 있다면 G = (2 Σ(i·x_i) - (n+1) Σx_i) / (n Σx_i)
# 대신 우리는 LARGE 사용 (수식으로만 처리). 여기서는 SUMPRODUCT 기반 페어와이즈.
# |x_i - x_j| 합: SUMPRODUCT(ABS(D{start}:D{end} - TRANSPOSE(D{start}:D{end})))
# openpyxl에서 TRANSPOSE 배열수식 사용:
gini_formula = (
    f'=SUMPRODUCT(ABS(D{D_START}:D{D_END}-TRANSPOSE(D{D_START}:D{D_END})))'
    f'/(2*B{gini_row + 1}^2*B{gini_row + 2})'
)
ws.cell(row=gini_row + 3, column=1, value='지니계수 (pair-wise)').font = BOLD_FONT
ws.cell(row=gini_row + 3, column=2, value=gini_formula)
ws[f'B{gini_row + 3}'] = openpyxl.worksheet.formula.ArrayFormula(
    f'B{gini_row + 3}', gini_formula
)
ws.cell(row=gini_row + 3, column=2).number_format = FLOAT_FMT
ws.cell(row=gini_row + 3, column=3,
        value='(0=완전평등, 1=극단불평등. 0.3↑ = 뚜렷한 격차)').font = SUB_FONT

# 상·하위 5개 자치구 바차트
# 정렬해서 상위 5, 하위 5 표시용 보조 영역
top_row = gini_row + 6
ws.cell(row=top_row, column=1, value='상·하위 5개 자치구 (1인당 이용 기준)').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=top_row + 1, column=1, value='순위').font = BOLD_FONT
ws.cell(row=top_row + 1, column=2, value='자치구').font = BOLD_FONT
ws.cell(row=top_row + 1, column=3, value='1인당 이용').font = BOLD_FONT
style_header(ws, top_row + 1, 3)

# 데이터 정렬 (static — 수식으로 할 수도 있으나 단순화)
sorted_gus = sorted(gus, key=lambda g: gu_2025[g] / population.get(g, 1), reverse=True)
for i in range(5):
    r = top_row + 2 + i
    g = sorted_gus[i]
    ws.cell(row=r, column=1, value=f'Top {i + 1}')
    ws.cell(row=r, column=2, value=g)
    ws.cell(row=r, column=3, value=round(gu_2025[g] / population.get(g, 1), 2))
    ws.cell(row=r, column=3).number_format = FLOAT_FMT

ws.cell(row=top_row + 7, column=1, value='━').font = SUB_FONT
for i in range(5):
    r = top_row + 8 + i
    g = sorted_gus[-(5 - i)]
    ws.cell(row=r, column=1, value=f'Bot {5 - i}')
    ws.cell(row=r, column=2, value=g)
    ws.cell(row=r, column=3, value=round(gu_2025[g] / population.get(g, 1), 2))
    ws.cell(row=r, column=3).number_format = FLOAT_FMT

style_data_range(ws, top_row + 2, top_row + 12, 1, 3)

# 바차트: 25자치구 1인당 이용 (내림차순 정렬된 결과)
chart_d = BarChart()
chart_d.type = 'bar'
chart_d.title = '자치구별 1인당 따릉이 이용 (2025)'
chart_d.y_axis.title = '자치구'
chart_d.x_axis.title = '1인당 이용(회/년)'
chart_d.style = 10
chart_d.width = 18
chart_d.height = 14
data_d = Reference(ws, min_col=4, min_row=3, max_row=D_END)
cats_d = Reference(ws, min_col=1, min_row=D_START, max_row=D_END)
chart_d.add_data(data_d, titles_from_data=True)
chart_d.set_categories(cats_d)
chart_d.series[0].graphicalProperties.solidFill = GREEN
ws.add_chart(chart_d, f'H{D_START}')

for c in range(1, 7):
    ws.column_dimensions[get_column_letter(c)].width = 16

print('✓ District_Analysis 추가')

# ═══════════════════════════════════════════════════════════════
# Sheet 13: Unit_Economics
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Unit_Economics')
ws.sheet_properties.tabColor = BLUE

ws['A1'] = '단위경제 — 자전거 1대당 일회전율'
ws['A1'].font = TITLE_FONT
ws['A2'] = '월별 이용건수 / (운영대수 × 월 일수) — 대당 회전율 시계열'
ws['A2'].font = SUB_FONT

# 운영대수 추정 (공개자료 기반)
fleet_by_year = {
    2019: 25000, 2020: 37500, 2021: 40500, 2022: 43500,
    2023: 45000, 2024: 45000, 2025: 45000,
}
# 월별 일수
def days_in_month(ym):
    y, m = int(str(ym)[:4]), int(str(ym)[4:])
    if m in (1, 3, 5, 7, 8, 10, 12): return 31
    if m == 2: return 29 if (y % 4 == 0 and y % 100 != 0) or y % 400 == 0 else 28
    return 30

headers = ['연월', '이용건수', '운영대수(추정)', '일수', '대당 일회전율']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

UE_START = 4
for idx, m in enumerate(monthly):
    r = UE_START + idx
    y = m['연도']
    ym = m['연월']
    ws.cell(row=r, column=1, value=ym)
    ws.cell(row=r, column=2, value=m['이용건수'])
    ws.cell(row=r, column=3, value=fleet_by_year.get(y, 45000))
    ws.cell(row=r, column=4, value=days_in_month(ym))
    ws.cell(row=r, column=5, value=f'=B{r}/(C{r}*D{r})')
    ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.cell(row=r, column=3).number_format = NUM_FMT
    ws.cell(row=r, column=5).number_format = FLOAT_FMT

UE_END = UE_START + N - 1
style_data_range(ws, UE_START, UE_END, 1, 5)

# 요약 (연도별 평균 회전율, 2023 vs 2025)
sum_row = UE_END + 2
ws.cell(row=sum_row, column=1, value='연도별 평균 대당 일회전율').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=sum_row + 1, column=1, value='연도').font = BOLD_FONT
ws.cell(row=sum_row + 1, column=2, value='평균 대당 일회전율').font = BOLD_FONT
style_header(ws, sum_row + 1, 2)
for i, yr in enumerate(sorted(fleet_by_year)):
    r = sum_row + 2 + i
    ws.cell(row=r, column=1, value=yr)
    ws.cell(row=r, column=2, value=(
        f'=AVERAGEIFS(E{UE_START}:E{UE_END},A{UE_START}:A{UE_END},">="&({yr}*100+1),A{UE_START}:A{UE_END},"<="&({yr}*100+12))'
    ))
    ws.cell(row=r, column=2).number_format = FLOAT_FMT
style_data_range(ws, sum_row + 2, sum_row + 1 + len(fleet_by_year), 1, 2)

# 2023→2025 하락폭
drop_row = sum_row + 10
ws.cell(row=drop_row, column=1, value='2023→2025 대당 회전율 변화').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=drop_row + 1, column=1, value='2023 평균').font = BOLD_FONT
ws.cell(row=drop_row + 1, column=2, value=f'=B{sum_row + 2 + 4}')  # 2023은 6번째(idx=4)
ws.cell(row=drop_row + 2, column=1, value='2025 평균').font = BOLD_FONT
ws.cell(row=drop_row + 2, column=2, value=f'=B{sum_row + 2 + 6}')
ws.cell(row=drop_row + 3, column=1, value='변화율').font = BOLD_FONT
ws.cell(row=drop_row + 3, column=2, value=f'=B{drop_row + 2}/B{drop_row + 1}-1')
ws.cell(row=drop_row + 3, column=2).number_format = PCT_FMT

# 연간 적자/건당 원단위 (참고치)
ctx_row = drop_row + 6
ws.cell(row=ctx_row, column=1, value='재무 참조 (뉴시스·국민일보 인용)').font = Font(bold=True, size=12, color=NAVY)
ws.cell(row=ctx_row + 1, column=1, value='시간당 요금').font = BOLD_FONT
ws.cell(row=ctx_row + 1, column=2, value=1000)
ws.cell(row=ctx_row + 1, column=2).number_format = '"₩"#,##0'
ws.cell(row=ctx_row + 2, column=1, value='연간 적자 (추정)').font = BOLD_FONT
ws.cell(row=ctx_row + 2, column=2, value=10_000_000_000)
ws.cell(row=ctx_row + 2, column=2).number_format = '"₩"#,##0'
ws.cell(row=ctx_row + 3, column=1, value='건당 적자 원단위 (2025)').font = BOLD_FONT
ws.cell(row=ctx_row + 3, column=2,
        value=f'=B{ctx_row + 2}/SUMIFS(B{UE_START}:B{UE_END},A{UE_START}:A{UE_END},">=202501",A{UE_START}:A{UE_END},"<=202512")')
ws.cell(row=ctx_row + 3, column=2).number_format = '"₩"#,##0.0'

# 차트: 대당 회전율 시계열
chart_ue = LineChart()
chart_ue.title = '자전거 1대당 일회전율 추이 — 2023 피크 대비 2025 하락'
chart_ue.y_axis.title = '대당 일회전율'
chart_ue.x_axis.title = '연월'
chart_ue.style = 10
chart_ue.width = 28
chart_ue.height = 12
data_ue = Reference(ws, min_col=5, min_row=3, max_row=UE_END)
cats_ue = Reference(ws, min_col=1, min_row=UE_START, max_row=UE_END)
chart_ue.add_data(data_ue, titles_from_data=True)
chart_ue.set_categories(cats_ue)
chart_ue.series[0].graphicalProperties.line.solidFill = BLUE
chart_ue.series[0].graphicalProperties.line.width = 24000
ws.add_chart(chart_ue, f'G3')

for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 16

print('✓ Unit_Economics 추가')

# ═══════════════════════════════════════════════════════════════
# Sheet 14: Misu_Adjustment
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Misu_Adjustment')
ws.sheet_properties.tabColor = '808080'

ws['A1'] = '"미상" 안분 보정 — 연령별 CAGR 재산출'
ws['A1'].font = TITLE_FONT
ws['A2'] = '원본 CAGR vs 미상을 각 연령 비율로 안분한 보정 CAGR 병기'
ws['A2'].font = SUB_FONT

# age_yearly가 미상 없이 연령별 건수만 제공한다고 가정
# '미상'은 gender_yearly 구조에만 있을 수 있음. 검증:
has_misu_age = any('미상' in a for a in age_yearly)
print('  age_yearly 미상 포함:', has_misu_age)

# gender_yearly에서 연도별 미상 비율 참고
misu_pct_by_year = {}
for gy in gender_yearly:
    y = gy['연도']
    tot = gy.get('M', 0) + gy.get('F', 0) + gy.get('미상', 0)
    misu = gy.get('미상', 0)
    misu_pct_by_year[y] = misu / tot if tot > 0 else 0

age_groups = ['~10대', '20대', '30대', '40대', '50대', '60대', '70대이상']

# 연령별 '미상 안분 보정' 값: age_i_adj = age_i * (1 / (1 - misu_pct))
# 즉 연령별 집계가 '공시된 값'이면, 미상은 각 연령에 비율대로 흡수됐다 가정
ay_2019 = next((a for a in age_yearly if a['연도'] == 2019), {})
ay_2025 = next((a for a in age_yearly if a['연도'] == 2025), {})
misu19 = misu_pct_by_year.get(2019, 0)
misu25 = misu_pct_by_year.get(2025, 0)

headers = ['연령대', '2019 원본', '2025 원본', '원본 CAGR', '2019 보정', '2025 보정',
           '보정 CAGR', '차이(pp)']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

M_START = 4
for i, ag in enumerate(age_groups):
    r = M_START + i
    v19 = ay_2019.get(ag, 0)
    v25 = ay_2025.get(ag, 0)
    # 원본 CAGR
    ws.cell(row=r, column=1, value=ag)
    ws.cell(row=r, column=2, value=v19 if v19 > 0 else 1)
    ws.cell(row=r, column=3, value=v25 if v25 > 0 else 1)
    ws.cell(row=r, column=4, value=f'=IFERROR((C{r}/B{r})^(1/6)-1,0)')
    # 보정: 원본 / (1 - misu_pct)
    ws.cell(row=r, column=5, value=f'=B{r}/(1-{misu19})')
    ws.cell(row=r, column=6, value=f'=C{r}/(1-{misu25})')
    ws.cell(row=r, column=7, value=f'=IFERROR((F{r}/E{r})^(1/6)-1,0)')
    ws.cell(row=r, column=8, value=f'=(G{r}-D{r})*100')
    ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.cell(row=r, column=3).number_format = NUM_FMT
    ws.cell(row=r, column=4).number_format = PCT_FMT
    ws.cell(row=r, column=5).number_format = NUM_FMT
    ws.cell(row=r, column=6).number_format = NUM_FMT
    ws.cell(row=r, column=7).number_format = PCT_FMT
    ws.cell(row=r, column=8).number_format = FLOAT_FMT

M_END = M_START + len(age_groups) - 1
style_data_range(ws, M_START, M_END, 1, 8)

# 해설
note_row = M_END + 3
notes = [
    f'• 2019 미상 비율: {misu19 * 100:.1f}%  /  2025 미상 비율: {misu25 * 100:.1f}%',
    '• 보정 가정: 미상(미등록) 이용자의 연령 분포가 등록 이용자와 동일',
    '• 미상 비율이 급감(57%→25%)했으므로 보정 전 CAGR은 **집계 기준 변화**에 의해 과대계상됨',
    '• 보정 후 CAGR은 연령별 실제 수요 증가율에 더 가까움',
    '• 특히 고성장으로 보이는 60대+의 CAGR이 보정 후 얼마나 축소되는지 확인',
]
ws.cell(row=note_row, column=1, value='해석 메모').font = Font(bold=True, size=12, color=NAVY)
for i, n in enumerate(notes):
    ws.cell(row=note_row + 1 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=note_row + 1 + i, start_column=1,
                   end_row=note_row + 1 + i, end_column=8)

# 차트: 보정 전후 CAGR 비교
chart_m = BarChart()
chart_m.title = '연령별 CAGR — 원본 vs 미상 보정'
chart_m.y_axis.title = 'CAGR'
chart_m.x_axis.title = '연령대'
chart_m.style = 10
chart_m.width = 22
chart_m.height = 12
# 두 시리즈: 원본(col 4) + 보정(col 7)
# add_data는 연속 범위를 기대하므로 둘을 개별 추가
data1 = Reference(ws, min_col=4, min_row=3, max_row=M_END)
chart_m.add_data(data1, titles_from_data=True)
data2 = Reference(ws, min_col=7, min_row=3, max_row=M_END)
chart_m.add_data(data2, titles_from_data=True)
cats_m = Reference(ws, min_col=1, min_row=M_START, max_row=M_END)
chart_m.set_categories(cats_m)
chart_m.series[0].graphicalProperties.solidFill = '999999'
chart_m.series[1].graphicalProperties.solidFill = PURPLE
ws.add_chart(chart_m, f'J{M_START}')

for c in range(1, 9):
    ws.column_dimensions[get_column_letter(c)].width = 15

print('✓ Misu_Adjustment 추가')

# ═══════════════════════════════════════════════════════════════
# Save and update meta
# ═══════════════════════════════════════════════════════════════
meta['v2_sheets'] = {
    'Residual_Analysis': {'start': RES_START, 'end': RES_END, 'stat_row': stat_row},
    'LINEST_Multi': {'x_start': X_START, 'x_end': X_END, 'linest_block_row': linest_block_row},
    'Forecast_Benchmark': {'start': FB_START, 'end': FB_END, 'summary_row': summary_row, 'fc_row': fc_row},
    'District_Analysis': {'start': D_START, 'end': D_END, 'gini_row': gini_row},
    'Unit_Economics': {'start': UE_START, 'end': UE_END, 'sum_row': sum_row},
    'Misu_Adjustment': {'start': M_START, 'end': M_END,
                        'misu19_pct': round(misu19 * 100, 1),
                        'misu25_pct': round(misu25 * 100, 1)},
}
meta['sheet_names'] = wb.sheetnames
with open('excel_meta.json', 'w') as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

wb.save(XLSX)
print(f"\n✅ V2 시트 추가 완료. 총 시트: {wb.sheetnames}")
print(f"   meta 업데이트: excel_meta.json")
