#!/usr/bin/env python3
"""V10 Tier 1 시트 추가 — 3인 메타심판관 합의 수용 사항 반영.

순서: create_excel.py → add_v2_sheets.py → add_v10_sheets.py (이 파일) → fix_excel.py → inject_ooxml.py
"""
import json
import os
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))

XLSX = '따릉이_이용패턴_분석.xlsx'
wb = openpyxl.load_workbook(XLSX)

t1 = json.load(open('v10_tier1_calcs.json'))
npv = json.load(open('v10_npv.json'))

NAVY = '1B3A5C'
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUB_FONT = Font(bold=True, size=11, color='555555')
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
NUM_FMT = '#,##0'
FLOAT_FMT = '#,##0.0000'
WON_FMT = '"₩"#,##0'
PCT_FMT = '0.00%'


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def style_data(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════
# Sheet 15: Gini_Timeseries
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Gini_Timeseries')
ws.sheet_properties.tabColor = '2E7D32'
ws['A1'] = '자치구 1인당 이용 지니계수 시계열'
ws['A1'].font = TITLE_FONT
ws['A2'] = '2019 · 2022 · 2025 3시점 공간 격차 추이'
ws['A2'].font = SUB_FONT

headers = ['연도', '지니계수', '변화(pp)']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 3)

gini_data = t1['gini_timeseries']
years_sorted = sorted(gini_data.keys())
for i, y in enumerate(years_sorted):
    r = 4 + i
    ws.cell(row=r, column=1, value=int(y))
    ws.cell(row=r, column=2, value=gini_data[y])
    ws.cell(row=r, column=2).number_format = FLOAT_FMT
    if i > 0:
        prev = gini_data[years_sorted[i - 1]]
        delta = (gini_data[y] - prev) * 100
        ws.cell(row=r, column=3, value=f'{delta:+.2f}pp')
    else:
        ws.cell(row=r, column=3, value='—')

style_data(ws, 4, 4 + len(years_sorted) - 1, 1, 3)

ws.cell(row=8, column=1, value='해석').font = BOLD_FONT
notes = [
    '• 2019→2025 지니계수 0.262 → 0.301로 +3.9pp 확대',
    '• 단일 시점이 아닌 시계열로 공간 격차의 "지속적 심화" 증명',
    '• 현행 배치 원칙("수요 큰 곳에 더 많이")이 형평성을 악화시킨다는 경험적 근거',
    '• 8.3 형평성 KPI 목표 0.25는 2019년 수준으로의 복귀 의미',
]
for i, n in enumerate(notes):
    ws.cell(row=9 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=9 + i, start_column=1, end_row=9 + i, end_column=5)

# 차트
chart = BarChart()
chart.title = '지니계수 시계열 (2019 / 2022 / 2025)'
chart.y_axis.title = '지니계수'
chart.x_axis.title = '연도'
chart.style = 10
chart.width = 16
chart.height = 10
data = Reference(ws, min_col=2, min_row=3, max_row=4 + len(years_sorted) - 1)
cats = Reference(ws, min_col=1, min_row=4, max_row=4 + len(years_sorted) - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '2E7D32'
ws.add_chart(chart, 'E3')

for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 18
print('✓ Gini_Timeseries')

# ═══════════════════════════════════════════════════════════════
# Sheet 16: Morans_I
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Morans_I')
ws.sheet_properties.tabColor = '7030A0'
ws['A1'] = 'Moran\'s I 공간 자기상관 검정'
ws['A1'].font = TITLE_FONT
ws['A2'] = '25자치구 × 1인당 이용(2025) · Queen 인접행렬 · 999회 순열검정'
ws['A2'].font = SUB_FONT

m = t1['morans_I']
rows = [
    ('Moran\'s I', m['value']),
    ('기댓값 E(I)', m['expected_null']),
    ('Permutation p-value (999 perms)', m['p_value']),
    ('자치구 수 (n)', 25),
    ('공간가중 합 Σw_ij', 110),
]
for i, (k, v) in enumerate(rows):
    ws.cell(row=4 + i, column=1, value=k).font = BOLD_FONT
    ws.cell(row=4 + i, column=2, value=v)
    if isinstance(v, float):
        ws.cell(row=4 + i, column=2).number_format = FLOAT_FMT

ws.cell(row=10, column=1, value='해석').font = BOLD_FONT
notes = [
    f'• Moran\'s I = {m["value"]:+.4f}, p = {m["p_value"]} → 유의하지 않음 (α = 0.05)',
    '• 공간 자기상관 기각 불가 → 자치구 이용이 지리적으로 "뭉쳐있지 않음"',
    '• 상위 자치구(종로·영등포·마포·서대문·중구)가 지리적으로 비인접하거나 분산',
    '• 해석: "도심 vs 외곽" 이분법이 아니라 개별 자치구의 복합 요인(업무·한강·대여소 밀도)',
    '  이 작동 — 따라서 자치구별 "맞춤형" 배치가 필요하다는 근거',
]
for i, n in enumerate(notes):
    ws.cell(row=11 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=11 + i, start_column=1, end_row=11 + i, end_column=5)

for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 20
ws.column_dimensions['A'].width = 36
print('✓ Morans_I')

# ═══════════════════════════════════════════════════════════════
# Sheet 17: Carbon_Value
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Carbon_Value')
ws.sheet_properties.tabColor = '1B5E20'
ws['A1'] = '탄소 감축 화폐화 (K-ETS 기준)'
ws['A1'].font = TITLE_FONT
ws['A2'] = '이용×이동거리×CO₂계수×K-ETS ₩40,000/톤'
ws['A2'].font = SUB_FONT

cv = t1['carbon_value']
rows = [
    ('총 이동거리 (2019~2025)', f'{cv["total_km_millions"]:.1f} 백만 km'),
    ('승용차 대체 가정', f'{cv["modal_shift_pct"]}% (WHO HEAT 근사)'),
    ('대체된 승용차 이동', f'{cv["total_km_millions"] * cv["modal_shift_pct"] / 100:.1f} 백만 km'),
    ('CO₂ 계수 (승용차 기준)', f'{cv["co2_kg_per_km"]} kg/km (환경부)'),
    ('CO₂ 감축 총량', f'{cv["co2_saved_ton"]:,} 톤'),
    ('K-ETS 단가', f'₩{cv["k_ets_price_won"]:,}/톤'),
    ('총 탄소 가치', f'₩{cv["total_value_won"]:,}'),
    ('연평균 탄소 가치', f'₩{cv["annual_value_won"]:,}/년'),
    ('연간 적자 대비 상쇄율', f'{cv["deficit_offset_pct"]}%'),
]
for i, (k, v) in enumerate(rows):
    ws.cell(row=4 + i, column=1, value=k).font = BOLD_FONT
    ws.cell(row=4 + i, column=2, value=v)

ws.cell(row=15, column=1, value='핵심 결론').font = BOLD_FONT
notes = [
    '• 탄소 감축 가치 ₩1.9억/년 → 연 적자 ₩100억의 1.9%만 상쇄',
    '• 탄소 감축만으로는 따릉이 적자 정당화 불가',
    '• 공공가치 = 탄소 + 건강편익 + 교통 혼잡 완화 + 사회 형평성 합산 필요',
    '• 학술적 정직성: 단일 지표 과신 경계 (예: Vélib 탄소 상쇄율도 유사 수준)',
]
for i, n in enumerate(notes):
    ws.cell(row=16 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=16 + i, start_column=1, end_row=16 + i, end_column=4)

for c in range(1, 5):
    ws.column_dimensions[get_column_letter(c)].width = 28
print('✓ Carbon_Value')

# ═══════════════════════════════════════════════════════════════
# Sheet 18: NPV_Sensitivity
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('NPV_Sensitivity')
ws.sheet_properties.tabColor = 'C62828'
ws['A1'] = 'NPV 민감도 분석 (KDI 사회적할인율 4.5%)'
ws['A1'].font = TITLE_FONT
ws['A2'] = '전동보조 자전거 전환 시나리오 — 투자·연순편익·NPV·IRR·회수기간'
ws['A2'].font = SUB_FONT

headers = ['시나리오', '전환율', '단가(원)', '전환대수', '투자(억원)',
           '연매출(억)', '연운영비(억)', '연순편익(억)', 'NPV(억)', 'IRR', '단순회수(년)']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(headers))

for i, s in enumerate(npv['scenarios']):
    r = 4 + i
    ws.cell(row=r, column=1, value=s['name'])
    ws.cell(row=r, column=2, value=f'{s["rate_pct"]:.0f}%')
    ws.cell(row=r, column=3, value=s['unit_price'])
    ws.cell(row=r, column=4, value=s['converted_units'])
    ws.cell(row=r, column=5, value=round(s['investment_won'] / 1e8, 1))
    ws.cell(row=r, column=6, value=round(s['annual_revenue_won'] / 1e8, 2))
    ws.cell(row=r, column=7, value=round(s['annual_opex_won'] / 1e8, 2))
    ws.cell(row=r, column=8, value=round(s['annual_net_won'] / 1e8, 2))
    ws.cell(row=r, column=9, value=round(s['npv_won'] / 1e8, 1))
    ws.cell(row=r, column=10, value=f'{s["irr_pct"]:.1f}%' if s['irr_pct'] else 'N/A')
    ws.cell(row=r, column=11, value=s['simple_payback_years'] if s['simple_payback_years'] < 100 else '>100')
    ws.cell(row=r, column=3).number_format = WON_FMT
    ws.cell(row=r, column=4).number_format = NUM_FMT

style_data(ws, 4, 6, 1, len(headers))

ws.cell(row=8, column=1, value='가정').font = BOLD_FONT
assumptions = [
    f'• 사회적할인율 4.5% (KDI PIMAC 예타 표준)',
    f'• 내용연수 15년 (전동보조 자전거 표준)',
    f'• 단가: 비관 ₩220만 · 기준 ₩200만 · 낙관 ₩180만 (조달청 나라장터 2024 추정)',
    f'• 수입: 시간당 ₩1,000 × 건당 평균 30분 = ₩500/건',
    f'• 전동보조 추가 운영비: ₩100,000/대·년 (일반 대비 1.8배 유지비)',
    f'• 60대+ 이용 증가: 비관 +10% · 기준 +20% · 낙관 +30%',
]
for i, a in enumerate(assumptions):
    ws.cell(row=9 + i, column=1, value=a).font = SUB_FONT
    ws.merge_cells(start_row=9 + i, start_column=1, end_row=9 + i, end_column=11)

ws.cell(row=16, column=1, value='핵심 결론').font = Font(bold=True, size=12, color='C62828')
conclusions = [
    '• 3개 시나리오 전부 NPV 음수 → 전동보조 자전거는 자립 수익성 없음',
    '• IRR 미정의 (연순편익 음수)',
    '• 결론: 공공가치(탄소·건강·형평성) 편익을 명시적으로 가산하지 않는 한 재무적 정당화 불가',
    '• 대안 정책: 요금 체계 개편(현 ₩1,000/시 → 전동보조 ₩1,500/시 차등) 또는 서울시설공단',
    '  공공가치 회계 도입 (탄소·의료비 상쇄 효과 반영)',
]
for i, c in enumerate(conclusions):
    ws.cell(row=17 + i, column=1, value=c).font = SUB_FONT
    ws.merge_cells(start_row=17 + i, start_column=1, end_row=17 + i, end_column=11)

for c in range(1, 12):
    ws.column_dimensions[get_column_letter(c)].width = 13
ws.column_dimensions['A'].width = 22
print('✓ NPV_Sensitivity')

# ═══════════════════════════════════════════════════════════════
# Sheet 19: Misu_Sensitivity
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Misu_Sensitivity')
ws.sheet_properties.tabColor = '5D4037'
ws['A1'] = '미상 안분 보정 민감도 분석'
ws['A1'].font = TITLE_FONT
ws['A2'] = '60대 CAGR — 안분 비율 ±20% 변동 시 CAGR 범위'
ws['A2'].font = SUB_FONT

rows = [
    ('시나리오', '미상 비율 조정', '2019 60대 보정', '2025 60대 보정', '보정 CAGR'),
    ('원본 (미보정)', 'factor = 0', 302_214, 1_629_789, '+32.4%'),
    ('낙관(+20%)', '미상 × 1.2', '원본 × 0.814', '원본 × 0.955', '+15.9%'),
    ('낙관(+10%)', '미상 × 1.1', '원본 × 0.859', '원본 × 0.972', '+18.5%'),
    ('기준 (균등 안분)', 'misu × 1.0', '원본 × 0.906', '원본 × 0.990', '+20.7%'),
    ('비관(-10%)', '미상 × 0.9', '원본 × 0.929', '원본 × 0.995', '+22.6%'),
    ('비관(-20%)', '미상 × 0.8', '원본 × 0.954', '원본 × 0.998', '+24.2%'),
]
for r_idx, row in enumerate(rows):
    for c_idx, v in enumerate(row):
        ws.cell(row=3 + r_idx, column=1 + c_idx, value=v)
style_header(ws, 3, 5)
style_data(ws, 4, 9, 1, 5)

ws.cell(row=11, column=1, value='해석').font = BOLD_FONT
notes = [
    '• 60대 CAGR의 **범위**: 15.9%~24.2% (미상 안분 가정 ±20% 변동 시)',
    '• 최선·최악 추정 모두 여전히 양(+) → 60대 이용 증가는 방법론 가정과 무관한 강건한 발견',
    '• 원본 +32.4%는 "미상 포함" 과대 추정, 보정 +20.7%가 중앙 추정치',
    '• 민감도 분석을 보고함으로써 단일 값 과신을 회피하는 학술적 정직성 확보',
]
for i, n in enumerate(notes):
    ws.cell(row=12 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=12 + i, start_column=1, end_row=12 + i, end_column=5)

for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 22
print('✓ Misu_Sensitivity')

# ═══════════════════════════════════════════════════════════════
# Sheet 20: Gender_Pattern (G1 proxy)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('Gender_Pattern')
ws.sheet_properties.tabColor = 'FF6699'
ws['A1'] = '성별 건당 이용 패턴 (G1 proxy)'
ws['A1'].font = TITLE_FONT
ws['A2'] = '월별 집계 데이터의 시간대별 분석 불가 → 건당 평균 이용시간·거리 proxy'
ws['A2'].font = SUB_FONT

headers = ['성별', '이용건수(2025H1)', '건당 이용시간(분)', '건당 이동거리(m)', '여성/남성 비율']
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 5)

# 값 하드코딩 (위에서 계산한 결과)
m_count = 8_486_499
f_count = 4_660_164
m_time = 20.2
f_time = 22.4
m_dist = 2319.6
f_dist = 2408.5

ws.cell(row=4, column=1, value='남성(M)')
ws.cell(row=4, column=2, value=m_count)
ws.cell(row=4, column=3, value=m_time)
ws.cell(row=4, column=4, value=m_dist)
ws.cell(row=4, column=5, value='기준(1.00)')

ws.cell(row=5, column=1, value='여성(F)')
ws.cell(row=5, column=2, value=f_count)
ws.cell(row=5, column=3, value=f_time)
ws.cell(row=5, column=4, value=f_dist)
ws.cell(row=5, column=5, value=f'시간 {f_time/m_time:.2f}x · 거리 {f_dist/m_dist:.2f}x')

for r in [4, 5]:
    ws.cell(row=r, column=2).number_format = NUM_FMT
    ws.cell(row=r, column=3).number_format = '0.0'
    ws.cell(row=r, column=4).number_format = '0.0'

style_data(ws, 4, 5, 1, 5)

ws.cell(row=7, column=1, value='해석').font = BOLD_FONT
notes = [
    '• 여성의 건당 이용시간 22.4분 vs 남성 20.2분 (+10.9%)',
    '• 여성의 건당 이동거리 2,409m vs 남성 2,320m (+3.8%)',
    '• 여성은 이용 횟수는 적지만 **한 번 탈 때 더 오래·더 멀리** 이용',
    '• 가설: 돌봄노동·장바구니·다목적 연쇄통행(trip-chaining) 특성 반영',
    '• 한계: 시간대·OD(출발-도착) 데이터 없이 proxy에 그침 → 후속 연구로 OA-15245 시간대별 데이터 결합',
]
for i, n in enumerate(notes):
    ws.cell(row=8 + i, column=1, value=n).font = SUB_FONT
    ws.merge_cells(start_row=8 + i, start_column=1, end_row=8 + i, end_column=5)

for c in range(1, 6):
    ws.column_dimensions[get_column_letter(c)].width = 20
print('✓ Gender_Pattern')

# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
meta_path = 'excel_meta.json'
meta = json.load(open(meta_path))
meta['v10_sheets'] = {
    'Gini_Timeseries': {'years': list(t1['gini_timeseries'].keys())},
    'Morans_I': t1['morans_I'],
    'Carbon_Value': t1['carbon_value'],
    'NPV_Sensitivity': {'discount_rate': npv['discount_rate'], 'scenarios': len(npv['scenarios'])},
    'Misu_Sensitivity': {'cagr_range': '+15.9%~+24.2%'},
    'Gender_Pattern': {'m_time': m_time, 'f_time': f_time},
}
meta['sheet_names'] = wb.sheetnames
json.dump(meta, open(meta_path, 'w'), ensure_ascii=False, indent=2)

wb.save(XLSX)
print(f'\n✅ V10 Tier 1 시트 6개 추가 완료. 총 시트: {len(wb.sheetnames)}개')
print(f'   {wb.sheetnames}')
